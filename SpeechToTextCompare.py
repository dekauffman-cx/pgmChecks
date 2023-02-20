#!/usr/bin/env python3
import speech_recognition as sr
from openpyxl import load_workbook
from pathlib import Path
import os.path as op
import wavio
#import SpeechObject
import os
import sys
import glob as gb
import xlsxwriter as xw
from fuzzywuzzy import fuzz   #uses python-Levenshtein module
from fuzzywuzzy import process #need to install the above module along with fuzzyywuzzy

# sketching out how this script will work 
# open prompt list with openpyxl
# Pick up the list of voice slots from the voice directory (stored locally)
# match each individual voice slot to a line in the prompt list 
#   1.  if it exists  read tte voice file intu the  recognizer and print out the speech to text 
#       compare the speech to text to the text in the prompt list to ensure accuracy 
#           (figured out how to do fuzzy matching score so we don't need an exact match -- fuzzy wuzzy will work --- low score == low match)
#   2.  if it doesn't exist in the prompt list Or the prompt list has a voice slot that is unaccounted for in the voice file directory 
#       mark it as an error and proceed to the next voice slot 
#   3.  excel file to hold the information.

def ExcelWrite(voiceFilePath, vDict):
    """creating Excel file to store the information """
    try:
        os.chdir(voiceFilePath)
        bn = op.basename(voiceFilePath)
        print("path: " + voiceFilePath)
        print("basename: " + op.basename(voiceFilePath))
        with xw.Workbook(voiceFilePath +'/PCH_' + bn + '.xlsx') as wb:
            header = wb.add_format({'bold': 1, "text_wrap": True})
            wrap = wb.add_format({'text_wrap': True})
            under_65 = wb.add_format({"text_wrap": True, "bg_color": "yellow"})
            listKeys = sorted(vDict)
            sheets = []
            for i in listKeys:
                keyArr = i.split('/')
                basename = keyArr[0]
                voice = keyArr[1]
                if basename not in sheets:
                    ws = wb.add_worksheet(basename)
                    sheets.append(basename)
                    ws.activate()
                    ws.freeze_panes(1, 0)
                    ws.write(0,0,"Voice File Name", header)
                    ws.set_column(0,0,30)
                    ws.write(0,1, 'Text from PromptList', header )
                    ws.set_column(1,2,40)
                    ws.write(0,2, "Text from Audio", header)
                    ws.write(0,3, "Fuzzy Ratio", header)
                    ws.set_column(3, 4, 8) 
                    ws.write(0,4, "Fuzzy Partial Ratio", header)
                    row = 1
                innerDict = list(vDict[i].values())
                textOfAUdio  = innerDict[0]
                textFromFile = innerDict[1]
                ratio = innerDict[2]
                partialRatio = innerDict[3]
                if ratio < 65:
                    ws.write(row, 0, voice, under_65)
                    ws.write(row, 1, textFromFile,under_65)
                    ws.write(row, 2, textOfAUdio, under_65)
                    ws.write(row, 3, ratio, under_65)
                    ws.write(row, 4, partialRatio, under_65)
                else:
                    ws.write(row, 0, voice)
                    ws.write(row, 1, textFromFile, wrap)
                    ws.write(row, 2, textOfAUdio, wrap)
                    ws.write(row, 3, ratio)
                    ws.write(row, 4, partialRatio)
                row += 1

    except: 
        print(sys.exc_info())

def Compare(textOfVoice, textFromFile): 
    try:
        ratio = fuzz.ratio(textOfVoice, textFromFile)
        part_ratio = fuzz.partial_ratio(textOfVoice, textFromFile)
        return ratio, part_ratio
    except:
        print(sys.exc_info())
        
def GetRecognition(path, wav, textFromFile, vDict):
    try:
        d = {}
        tr = sr.Recognizer()
        combPathWav = path + '/' + wav
        if op.isfile(wav) == True:
            temp = 'temp.' + wav
            os.system("touch " + temp)
            cmd = '/usr/bin/sox ' + wav + ' -e signed-integer ' + temp
            os.system(cmd)
            p = '777'
            os.chmod(temp, int(p, base = 8))
            with sr.AudioFile(temp) as source:
                audio_data = tr.record(source)
                textOfAudio = tr.recognize_google(audio_data, language='en-US')
            (ratio, pr) = Compare(textOfAudio, textFromFile)
            os.remove(temp)
            if combPathWav in vDict:
                print("this shouldn't happen With all unique keys")
            else:
                d = {"TextOfAudio": textOfAudio, "TextFromFile": textFromFile, "ratio": ratio, "partial_ratio": pr }
                vDict[combPathWav] = d
        else:
            textOfAudio = "Voice File does not exist"
            d = {"TextOfAudio": textOfAudio, "TextFromFile": textFromFile, "ratio": 0, "partial_ratio": 0 }
            vDict[combPathWav] = d
    except UnboundLocalError: 
        tb = sys.exc_info()
        message = "check the content of this file"
        d = {"TextOfAudio": message, "TextFromFile": textFromFile, "ratio": 0, "partial_ratio": 0 }
        vDict[combPathWav] = d
    except sr.UnknownValueError: 
        tb = sys.exc_info()
        message = "check the content of this file"
        d = {"TextOfAudio": message, "TextFromFile": textFromFile, "ratio": 0, "partial_ratio": 0 }
        vDict[combPathWav] = d
    except FileNotFoundError:
        textOfAudio = "Voice File does not exist"
        d = {"TextOfAudio": textOfAudio, "TextFromFile": textFromFile, "ratio": 0, "partial_ratio": 0 }
        vDict[combPathWav] = d


def LoadExcel(ef, voiceDict, missingFiles):
    print(ef)
    try:
        dirArr = gb.glob("*")
        dirArr =  sorted(dirArr)
        print(dirArr)
        wb = load_workbook(filename = ef, read_only = True)
        sheet_obj = wb.active
        numRows = sheet_obj.max_row
        rp = sheet_obj.cell(3, 1).value
        os.chdir(rp)
        listVoice = os.listdir('.')
        print(listVoice)
        for i in range(3, numRows):
            rp = sheet_obj.cell(i, 1).value
            rVoice = sheet_obj.cell(i, 2).value
            wavFile = rVoice.strip() + '.wav'
            path = os.getcwd()
            print(i, numRows, path, rp, rVoice)
            rowText = sheet_obj.cell(i, 3).value
            if op.basename(path) != rp.strip():
                print(rp)
                os.chdir("..")
                os.chdir(rp.rstrip())
                print(op.basename(path))
                listVoice = os.listdir('.')
                print(listVoice)
            print(rp, op.basename(path))
            if wavFile not in listVoice:
               missingFiles.append(wavFile)
            GetRecognition(rp, wavFile, rowText, voiceDict)
        wb.close()
    except:
        tb = sys.exc_info()
        print(tb)
        print(rVoice)
        print(op.basename(path))

def do_CheckPaths(vp, ef):
    """checking paths to make sure they exist """
 
    if op.isfile(ef) == False:
        print("File not found -- check the name")
        sys.exit(1)
    else: 
        print("File found: "+ ef)
    
    if op.isabs(vp) == False:
        print("Voice path not found -- check the path")
        sys.exit(1)
    else:
        os.chdir(vp)
        print("Voice directory found: " + vp)


if __name__ == "__main__":
    """ do checks on voice file path and xlsx file """
    if len(sys.argv) == 3:
        voiceFilePath = sys.argv[1] 
        excelFile = sys.argv[2]
        do_CheckPaths(voiceFilePath, excelFile)
        voiceDict = {}
        missingFiles = []
        LoadExcel(excelFile, voiceDict, missingFiles)
        ExcelWrite(voiceFilePath,  voiceDict)
    else:
        print("required number of args has to be two")
        print("must have voice file path and excel file name plus directory in that order")
        sys.exit(0)

